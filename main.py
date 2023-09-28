import os
import sys
import time
import datetime

import win32api
import win32con

from openpyxl import Workbook
from PySide6.QtWidgets import QMainWindow, QApplication, QDialog, QAbstractItemView, QSystemTrayIcon, QMenu, QInputDialog, QFileDialog, QMessageBox
from PySide6.QtGui import QStandardItemModel, QStandardItem, QIcon, QAction
from PySide6.QtSql import QSqlTableModel
from PySide6.QtCore import Qt, QEvent, QSettings


from ui_window import Ui_MainWindow
from ui_dialog import Ui_Dialog
from db import SQLite3DB
from qcounter import KeyboardThread, MouseThread

import app_res


class CounterDialog(QDialog, Ui_Dialog):
    def __init__(self, parent=None, count_type='month'):
        self.parent = parent
        self.count_type = count_type
        super().__init__(parent)
        self.setupUi(self)
        count_type_name = '月度' if self.count_type=='month' else '年度'
        self.setWindowTitle(count_type_name + '按键数量统计表')

        model = QSqlTableModel(self.tableView, parent.cur_db.db)
        if count_type == 'month':
            cur_sql = parent.cur_db.get_cur_month_count_sql()
        else:
            cur_sql = parent.cur_db.get_cur_year_count_sql()
        print(cur_sql)
        model.setQuery(cur_sql)

        model.setHeaderData(model.fieldIndex("code"), Qt.Horizontal, self.tr("键值"))
        model.setHeaderData(model.fieldIndex("name"), Qt.Horizontal, self.tr("键名"))
        model.setHeaderData(model.fieldIndex("is_key_val"), Qt.Horizontal, self.tr("类型"))
        model.setHeaderData(model.fieldIndex("counter"), Qt.Horizontal, self.tr("数量"))

        self.tableView.setModel(model)
        self.tableView.setColumnHidden(model.fieldIndex("is_key"), True)

        self.tableView.setSelectionMode(QAbstractItemView.SingleSelection)
        self.tableView.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.tableView.resizeColumnsToContents()

        self.select_data = []
        if count_type == 'month':
            self.select_data = parent.cur_db.get_all_month()
        else:
            self.select_data = parent.cur_db.get_all_year()
        
        self.comboBox.clear()
        for item in self.select_data:
            self.comboBox.addItem(item['name'])
        self.comboBox.currentIndexChanged.connect(self.selectionChange)

        self.pushButton.clicked.connect(self.export_data)

    
    def selectionChange(self, i):
        print('select  %d item value: %s' % (i, self.comboBox.currentText()))
        cur_item = self.select_data[i]
        model = QSqlTableModel(self.tableView, self.parent.cur_db.db)
        if self.count_type == 'month':
            cur_sql = self.parent.cur_db.get_month_count_sql(month=cur_item['code'])
        else:
            cur_sql = self.parent.cur_db.get_year_count_sql(year=cur_item['code'])
        print(cur_sql)
        model.setQuery(cur_sql)

        model.setHeaderData(model.fieldIndex("code"), Qt.Horizontal, self.tr("键值"))
        model.setHeaderData(model.fieldIndex("name"), Qt.Horizontal, self.tr("键名"))
        model.setHeaderData(model.fieldIndex("is_key_val"), Qt.Horizontal, self.tr("类型"))
        model.setHeaderData(model.fieldIndex("counter"), Qt.Horizontal, self.tr("数量"))

        self.tableView.setModel(model)
        self.tableView.setColumnHidden(model.fieldIndex("is_key"), True)

        self.tableView.resizeColumnsToContents()

    def export_data(self):
        cur_path = "%s%s%s%s" % (os.getcwd(), os.sep, self.comboBox.currentText(),'键盘鼠标使用情况.xlsx')
        print(cur_path)
        filepath = QFileDialog.getSaveFileName(caption='导出统计数据到本电脑', dir=cur_path, filter='xlsx(*.xlsx)')
        print(filepath)

        cur_item = self.select_data[self.comboBox.currentIndex()]
        if self.count_type == 'month':
            data = self.parent.cur_db.get_month_count(month=cur_item['code'])
        else:
            data = self.parent.cur_db.get_year_count(year=cur_item['code'])
        
        wb = Workbook()
        ws = wb.active
        ws.cell(1, 1).value = '键值'
        ws.cell(1, 2).value = '键名'
        ws.cell(1, 3).value = '类型'
        ws.cell(1, 4).value = '数量'
        i = 2
        for row in data:
            print(row)
            print(row['name'])
            print(row['is_key'])
            print(type(row['is_key']))
            is_key = row['is_key']
            ws.cell(i, 1).value = row['code']
            ws.cell(i, 2).value = row['name']
            ws.cell(i, 3).value = row['is_key_val']
            ws.cell(i, 4).value = row['counter']
            i += 1

        wb.save(filepath[0])
        wb.close()
        QMessageBox.information(None, '导出成功提醒', '数据导出成功，请查看', QMessageBox.Ok)

        cur_path = os.path.dirname(os.path.abspath(filepath[0]))
        print(cur_path)
        os.system("start explorer %s" % cur_path)



class KeyMouCounterWindow(QMainWindow):
    def __init__(self, debug=False, keymouse_debug=False, db_debug=False):
            super(KeyMouCounterWindow, self).__init__()
            self.ui = Ui_MainWindow()
            self.ui.setupUi(self)
            
            self.setWindowTitle('键盘鼠标计使用统计小助手')
            self.setWindowIcon(QIcon(':/favicon.png')) 
            self.name = 'KeyMouCheck'
            

            self.ui.lineEdit.setReadOnly(True)
            self.ui.lineEdit_2.setReadOnly(True)
            self.ui.lineEdit_3.setReadOnly(True)
            self.ui.lineEdit_4.setReadOnly(True)
            self.ui.lineEdit_5.setReadOnly(True)
            self.ui.lineEdit_6.setReadOnly(True)

            self.setWindowFlags(Qt.WindowMinimizeButtonHint|Qt.WindowCloseButtonHint)

            self.ui.pushButton.clicked.connect(self.month_counter_show)
            self.ui.pushButton_2.clicked.connect(self.year_counter_show)
            self.ui.checkBox.stateChanged.connect(self.auto_run_control)

            self.debug = debug
            self.keymouse_debug = keymouse_debug
            self.db_debug = db_debug
            
            self.cur_db = SQLite3DB(self.db_debug)
            self.is_counter = True
            self.can_closed = False


            self.ui.lineEdit.setReadOnly(True)
            
            self.ui.lineEdit_2.setReadOnly(True)
            self.ui.lineEdit_3.setReadOnly(True)
            self.ui.lineEdit_4.setReadOnly(True)
            self.ui.lineEdit_5.setReadOnly(True)
            self.ui.lineEdit_6.setReadOnly(True)

            self.ui.lineEdit.setFocusPolicy(Qt.NoFocus)
            self.ui.lineEdit_2.setFocusPolicy(Qt.NoFocus)
            self.ui.lineEdit_3.setFocusPolicy(Qt.NoFocus)
            self.ui.lineEdit_4.setFocusPolicy(Qt.NoFocus)
            self.ui.lineEdit_5.setFocusPolicy(Qt.NoFocus)
            self.ui.lineEdit_6.setFocusPolicy(Qt.NoFocus)

            self.day_keyboard_count = self.cur_db.get_day_keyboard_count()
            self.day_mouse_count = self.cur_db.get_day_mouse_count()
            self.month_keyboard_count = self.cur_db.get_month_keyboard_count()
            self.month_mouse_count =  self.cur_db.get_month_mouse_count()
            self.year_keyboard_count = self.cur_db.get_year_keyboard_count()
            self.year_mouse_count = self.cur_db.get_year_mouse_count()

            self.ui.lineEdit.setText(str(self.day_keyboard_count))
            self.ui.lineEdit_2.setText(str(self.month_keyboard_count))
            self.ui.lineEdit_3.setText(str(self.year_keyboard_count))
            self.ui.lineEdit_4.setText(str(self.day_mouse_count))
            self.ui.lineEdit_5.setText(str(self.month_mouse_count))
            self.ui.lineEdit_6.setText(str(self.year_mouse_count))

            

            self.keyboard_counter = KeyboardThread(window=self, debug=self.keymouse_debug)
            self.mouse_counter = MouseThread(window=self, debug=self.keymouse_debug)

            self.setting = QSettings('config.ini', QSettings.IniFormat)
            if self.setting.value('setting/auto_run') is None:
                self.setting.setValue('setting/auto_run', 1)
            if int(self.setting.value('setting/auto_run'))== 1:
                self.auto_run()


    def closeEvent(self, event):
        if self.can_closed :
            #time.sleep(1)
            
            self.keyboard_counter.stop()
            self.keyboard_counter.quit()
            while self.keyboard_counter.isRunning():
                exit_code = self.keyboard_counter.wait(1)
            if self.debug:
                print("Close Keyboard Counter Thread")
            self.mouse_counter.stop()
            self.mouse_counter.quit() 
            while self.mouse_counter.isRunning():
                exit_code = self.mouse_counter.wait(1)
            if self.debug:
                print("Close Mouse Counter Thread")
            if self.debug:
                print("Close Database Connection")
            self.cur_db.close()
            time.sleep(1)
            if self.debug:
                print("Quit Application")
            event.accept()
        else:
            event.ignore()
            self.setVisible(False)

    def changeEvent(self, event):
        if event.type() == QEvent.WindowStateChange:
            if self.windowState() & Qt.WindowMinimized:
                event.ignore()
                self.setVisible(False)
        super(KeyMouCounterWindow, self).changeEvent(event)

    def month_counter_show(self):
        self.counter_dialog =  CounterDialog(parent=self)
        self.counter_dialog.show()

    def year_counter_show(self):
        self.counter_dialog =  CounterDialog(parent=self,count_type='year')
        self.counter_dialog.show()


    def auto_run(self):
        app_file = os.path.realpath(sys.argv[0])
        app_filename = os.path.basename(app_file)
        filename, extension = os.path.splitext(app_filename)
        stat = False
        if self.debug:
            print('App Fielname %s, %s' % (app_file, app_filename)) 
            print('App File Extentsion %s - [%s]' % (filename, extension)) 
        if extension == '.exe':
            KeyName = 'HKEY_CURRENT_USER\\SOFTWARE\\Microsoft\\Windows\\CurrentVersion\\Run'
            name = self.name
            reg = QSettings(KeyName, QSettings.NativeFormat)
            #print(reg)
            if not reg.value(name):
                reg.setValue(name, app_file)
            if reg.value(name):
                self.ui.checkBox.setChecked(True)
                self.setting.setValue('setting/auto_run', 1)
            else:
                self.ui.checkBox.setChecked(False)

            if self.debug:
                print("Regedit Name %s value is %s" % (name, reg.value(name)))
            stat = True
        return stat 

    def auto_run_cancel(self):
        app_file = os.path.realpath(sys.argv[0])
        app_filename = os.path.basename(app_file)
        filename, extension = os.path.splitext(app_filename)
        stat = False
        if self.debug:
            print('App Fielname %s, %s' % (app_file, app_filename)) 
            print('App File Extentsion %s - [%s]' % (filename, extension)) 
        if extension == '.exe':
            KeyName = 'HKEY_CURRENT_USER\\SOFTWARE\\Microsoft\\Windows\\CurrentVersion\\Run'
            name = self.name
            reg = QSettings(KeyName, QSettings.NativeFormat)
            if self.debug:
                print("Regedit Name %s value is %s" % (name, reg.value(name)))
            if reg.value(name):
                reg.remove(name)
                self.setting.setValue('setting/auto_run', 0)
            stat = True
        return stat 
    
    def auto_run_control(self):
        if self.ui.checkBox.isChecked():
            self.auto_run()
        else:
            self.auto_run_cancel()


class Tray(QSystemTrayIcon):
    def __init__(self, ui):
        super(Tray,self).__init__()
        self.ui = ui  # 传入主程序
        self.debug = self.ui.debug
        self.setIcon(QIcon(':/favicon.png'))  # 托盘图标
        self.setToolTip('键盘鼠标计使用统计小助手')  # 鼠标点在图标上的时候显示的气泡提示
        self.activated.connect(self.clickedIcon)  # 点击信号
        self.menu()
        self.show()

    def clickedIcon(self,reason):
        if reason == QSystemTrayIcon.DoubleClick:   #2
            self.trayClickedEvent()
        elif reason == QSystemTrayIcon.Trigger:  #3
            self.contextMenu()

    def menu(self):
        menu = QMenu()
        action_1 = QAction('显示', self, triggered=self.show_window)
        menu.addAction(action_1)
        action_2 = QAction('退出', self, triggered=self.triggered)
        menu.addAction(action_2)
        self.setContextMenu(menu)

    # 单击托盘图标，程序在隐藏和显示之间来回切换
    def trayClickedEvent(self):
        if not self.ui.isVisible():
            if self.debug:
                print("显示")
            self.ui.setVisible(True)
            self.ui.raise_()
            self.ui.activateWindow()
        else:
            if self.debug:
                print("隐藏")
            self.ui.setVisible(False)

    def show_window(self):
        if self.debug:
            print("显示")
        self.ui.setVisible(True)
        self.ui.raise_()
        self.ui.activateWindow()

    def triggered(self):
        ok = True
        text = 'kill'
        #text, ok = QInputDialog.getText(self.ui, "应用退出对话框", "请输入退出系统密码：")
        if ok and text=='kill':
            if self.debug:
                print("关闭程序")

            self.deleteLater()  # 删除托盘图标，无此操作的话，程序退出后托盘图标不会自动清除
            
            self.ui.can_closed = True
            if not self.ui.isVisible():
                self.ui.setVisible(True)
                self.ui.raise_()
                self.ui.activateWindow()
   
            self.ui.close()  # 后面会重写closeEvent，所以这里换一个退出程序的命令           


if __name__ == "__main__":
    app = QApplication([])

    window = KeyMouCounterWindow(debug=True, keymouse_debug=False, db_debug=False)

    window.keyboard_counter.start()
    window.mouse_counter.start()

    tray = Tray(ui=window)

    #window.resize(800, 600)
    window.show()
    #dialog.run_inspect()

    sys.exit(app.exec())